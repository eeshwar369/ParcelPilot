from __future__ import annotations

from pathlib import Path
from typing import Any

from backend.app.data_store import DataStore


RAW_DIR = Path(__file__).resolve().parents[2] / "data" / "raw"


def ingest_available_files(store: DataStore) -> dict[str, Any]:
    """Best-effort ingestion hook.

    The assessment data pack uses PDFs and XLSX. This demo remains runnable with no
    dependencies, then upgrades automatically if optional libraries are installed.
    """
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    files = [path for path in RAW_DIR.iterdir() if path.is_file() and path.name != ".gitkeep"]
    if not files:
        return {"status": "demo_data", "message": "No raw data files found; using built-in demo data.", "files": []}

    store.state["documents"] = []
    store.state["accounts"] = []
    store.state["orders"] = []
    store.state["tickets"] = []
    store.state["pending_actions"] = []
    store.state["escalations"] = []
    store.state["audit_events"] = []
    store.state["model_usage_events"] = []

    loaded: list[str] = []
    skipped: list[str] = []
    for path in files:
        suffix = path.suffix.lower()
        if suffix in {".txt", ".md"}:
            text = path.read_text(encoding="utf-8", errors="ignore")
            store.state["documents"].append(
                {
                    "id": f"raw_{path.stem}",
                    "name": path.stem.replace("_", " "),
                    "source_type": "uploaded_document",
                    "authority": "uploaded_context",
                    "status": "current",
                    "account_id": None,
                    "page": 1,
                    "text": text,
                }
            )
            loaded.append(path.name)
        elif suffix == ".pdf":
            if _try_pdf(path, store):
                loaded.append(path.name)
            else:
                skipped.append(f"{path.name} (install pypdf for PDF extraction)")
        elif suffix == ".xlsx":
            if _try_xlsx(path, store):
                loaded.append(path.name)
            else:
                skipped.append(f"{path.name} (install openpyxl for workbook extraction)")
        else:
            skipped.append(path.name)

    store.save()
    return {"status": "ok", "loaded": loaded, "skipped": skipped}


def _try_pdf(path: Path, store: DataStore) -> bool:
    try:
        from pypdf import PdfReader  # type: ignore
    except Exception:
        return False
    reader = PdfReader(str(path))
    text_parts = []
    for i, page in enumerate(reader.pages, start=1):
        text_parts.append(f"[page {i}]\n{page.extract_text() or ''}")
    store.state["documents"].append(
        {
            "id": f"pdf_{path.stem}",
            "name": path.stem.replace("_", " "),
            "source_type": _source_type_from_name(path.name),
            "authority": _authority_from_name(path.name),
            "status": "deprecated" if "DEPRECATED" in path.name.upper() else "current",
            "account_id": _account_from_name(path.name),
            "page": 1,
            "text": "\n".join(text_parts),
        }
    )
    return True


def _try_xlsx(path: Path, store: DataStore) -> bool:
    try:
        import openpyxl  # type: ignore
    except Exception:
        return False
    workbook = openpyxl.load_workbook(path, data_only=True)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        if sheet_name.lower() == "readme":
            _load_readme(rows, store)
        elif sheet_name.lower() == "accounts":
            store.state["accounts"] = [_normalize_account(row) for row in _dict_rows(rows)]
        elif sheet_name.lower() == "orders":
            store.state["orders"] = [_normalize_order(row) for row in _dict_rows(rows)]
        elif sheet_name.lower() == "tickets":
            store.state["tickets"] = [_normalize_ticket(row) for row in _dict_rows(rows)]
    return True


def _dict_rows(rows: list[tuple]) -> list[dict]:
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    result = []
    for row in rows[1:]:
        if not any(value is not None for value in row):
            continue
        result.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return result


def _load_readme(rows: list[tuple], store: DataStore) -> None:
    for row in rows:
        if row and str(row[0]).strip().lower() == "dataset snapshot" and len(row) > 1 and row[1]:
            store.state["snapshot_time"] = str(row[1])


def _normalize_account(row: dict) -> dict:
    return {
        "id": row.get("account_id"),
        "name": row.get("account_name"),
        "tier": row.get("plan"),
        "status": row.get("status"),
        "csm": row.get("csm"),
        "contract_file": row.get("contract_file"),
        "premium_support": bool(row.get("premium_support")),
        "notes": row.get("notes"),
        "raw_data": row,
    }


def _normalize_order(row: dict) -> dict:
    return {
        "id": row.get("order_id"),
        "account_id": row.get("account_id"),
        "carrier": row.get("carrier"),
        "service_level": row.get("status"),
        "status": row.get("status"),
        "booked_at": _stringify_dt(row.get("booked_at")),
        "scheduled_pickup_at": _stringify_dt(row.get("pickup_window_end") or row.get("pickup_window_start")),
        "pickup_window_start": _stringify_dt(row.get("pickup_window_start")),
        "pickup_window_end": _stringify_dt(row.get("pickup_window_end")),
        "actual_pickup_at": _stringify_dt(row.get("pickup_actual_at")),
        "shipment_fee_inr": row.get("shipment_fee_inr"),
        "carrier_fault": bool(row.get("carrier_fault")),
        "customer_fault": bool(row.get("customer_fault")),
        "cancellation_requested_at": _stringify_dt(row.get("cancellation_requested_at")),
        "notes": row.get("notes"),
        "raw_data": row,
    }


def _normalize_ticket(row: dict) -> dict:
    subject = str(row.get("subject") or "")
    description = str(row.get("description") or "")
    return {
        "id": row.get("ticket_id"),
        "account_id": row.get("account_id"),
        "order_id": _extract_order_id(subject + " " + description),
        "category": _infer_category(subject, description),
        "severity": _infer_severity(subject, description),
        "status": row.get("status"),
        "created_at": _stringify_dt(row.get("created_at")),
        "last_updated_at": _stringify_dt(row.get("last_customer_message_at")),
        "summary": f"{subject}: {description}",
        "subject": subject,
        "description": description,
        "channel": row.get("channel"),
        "assigned_to": row.get("assigned_to"),
        "historical_resolution": row.get("historical_resolution"),
        "raw_data": row,
    }


def _stringify_dt(value) -> str | None:
    if value is None:
        return None
    return value.isoformat() if hasattr(value, "isoformat") else str(value)


def _extract_order_id(text: str) -> str | None:
    import re

    match = re.search(r"\bORD-\d+\b", text, flags=re.I)
    return match.group(0).upper() if match else None


def _infer_category(subject: str, description: str) -> str:
    text = f"{subject} {description}".lower()
    if "api key" in text or "exposure" in text:
        return "security"
    if "500" in text or "failing" in text:
        return "platform_outage"
    if "bulk upload" in text or "csv" in text:
        return "bulk_upload"
    if "billing" in text:
        return "billing_admin"
    if "booked" in text or "pickup" in text:
        return "tracking_status"
    return "support"


def _infer_severity(subject: str, description: str) -> str:
    text = f"{subject} {description}".lower()
    if "all shipment" in text or "api key exposure" in text or "production api key" in text:
        return "high"
    if "fails" in text or "pickup" in text or "500" in text:
        return "medium"
    return "low"


def _source_type_from_name(name: str) -> str:
    lowered = name.lower()
    if "agreement" in lowered:
        return "customer_agreement"
    if "sop" in lowered or "credit" in lowered or "cancellation" in lowered:
        return "sop"
    if "policy" in lowered:
        return "support_policy"
    return "uploaded_document"


def _authority_from_name(name: str) -> str:
    lowered = name.lower()
    if "agreement" in lowered:
        return "customer_contract"
    if "deprecated" in lowered:
        return "deprecated_policy"
    if "sop" in lowered or "credit" in lowered or "cancellation" in lowered:
        return "current_sop"
    if "policy" in lowered:
        return "current_policy"
    return "uploaded_context"


def _account_from_name(name: str) -> str | None:
    lowered = name.lower()
    if "northstar" in lowered:
        return "ACCT-001"
    if "lumenworks" in lowered:
        return "ACCT-002"
    return None
